import itertools

from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl import styles
from bit import *
from itertools import permutations
bits = 4
f = lambda A, B, C, D: A & B & C & D  # Função a ser calculada
nome = "testeand.xlsx"  # Nome do arquivo gerado em excel


def imprimematriz(mat, espaco=5):  # *Mostra* a matriz formatada
    for i in range(len(mat)):  # Percorre as linhas
        # print(i + 1, end=' ')  # Printa o número da linha
        for j in range(len(mat[i])):  # Percorre cada termo da linha
            if j < len(mat[i]) - 1:  # Enquanto j n for o ultimo termo da linha
                print(f'{mat[i][j]:<{espaco}}', end=' ')
            else:
                print(f'{mat[i][j]}')


def to_binario(decimal):
    bitsList = []

    while decimal > 0:
        bitsList = [decimal % 2] + bitsList
        decimal //= 2

    if len(bitsList) == 0:
        return [0]

    return bitsList


def save_excel(sheet, name):
    linha = coluna = -1
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=len(m[0]) + 1, max_row=len(m) + 1):
        linha += 1
        for cell in row:
            coluna += 1

            cell.value = m[linha][coluna]

        coluna = -1

    sheet.column_dimensions[get_column_letter(3)].width = 4
    sheet.column_dimensions[get_column_letter(bits + 4)].width = 4
    sheet.sheet_view.showGridLines = False

    # Pinta os espaços
    my_black = styles.colors.Color()
    my_fill = styles.fills.PatternFill(patternType='solid', fgColor=my_black)
    for c in sheet["C:C"]:
        c.fill = my_fill
    for c in sheet[f'{get_column_letter(bits + 4)}:{get_column_letter(bits + 4)}']:
        c.fill = my_fill
    # Alinha os valores
    # for cel in sheet:
    #    cel.alignment = styles.Alignment(vertical='center', horizontal='center')

    book.save(name)


def imprimekarnaught(mat):
    karnot = [[0] * bits for z in range(bits)] if bits % 2 == 0 else [[0] * (2) ** (bits // 2) for z in
                                                                      range(2 ** (bits // 2 + 1))]
    print(karnot)
    for c in m:
        print(c[-1])

    mapadicio = {}
    raise NotImplementedError


book = Workbook()
sheet = book.active
values = [Bit(0), Bit(1)]
m=list(itertools.product((0,1), repeat=bits)) # Matriz com as possibilidades de bits
r = list()  # Vetor Respostas

# Vetor Resultado
for item in m:
    r.append(f(*item))

print(m)

# imprimematriz(m)
#imprimekarnaught(m)
# Escreve no excel a tabela verdade
save_excel(sheet, nome)

print(list(map(str,r)))